#!/usr/bin/env python3
"""
xlsx-to-json.py — MasjidBox XLSX → prayer-times.json converter

Reads the masjidbox athan xlsx file and produces a single JSON file
containing all 365 days of adhan + iqamah times, keyed by DD/MM/YYYY.

Usage:
    python xlsx-to-json.py                          # looks for xlsx in parent dir
    python xlsx-to-json.py path/to/file.xlsx        # specify xlsx path

Outputs:
    assets/data/prayer-times.json      (for WordPress theme)
    ../js/prayer-times.json            (for static HTML prototype)

@package emc-theme
"""

import json
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# ── Configuration ──────────────────────────────────────────────────────────
ATHAN_SHEET    = "Athan with overwrite"
ATHAN_NO_OW    = "Athan without overwrite"
IQAMAH_SHEET   = "Iqamah"

# Expected column headers (lowercase) for validation
ATHAN_COLS  = ["date", "fajr", "sunrise", "dhuhr", "asr", "maghrib", "isha", "jumuah"]
IQAMAH_COLS = ["date", "fajr", "dhuhr", "asr", "maghrib", "isha", "jumuah"]

# Output paths (relative to this script's directory)
THEME_OUTPUT   = os.path.join("assets", "data", "prayer-times.json")
STATIC_OUTPUT  = os.path.join("..", "js", "prayer-times.json")


def find_xlsx(explicit_path=None):
    """Locate the masjidbox xlsx file."""
    if explicit_path and os.path.isfile(explicit_path):
        return explicit_path

    # Look in parent directory for any file matching the pattern
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)

    for f in os.listdir(parent_dir):
        if f.lower().startswith("masjidbox") and f.lower().endswith(".xlsx"):
            return os.path.join(parent_dir, f)

    # Also check the script's own directory
    for f in os.listdir(script_dir):
        if f.lower().startswith("masjidbox") and f.lower().endswith(".xlsx"):
            return os.path.join(script_dir, f)

    return None


def validate_time(time_str, context=""):
    """Validate a time string is in HH:MM format."""
    if not time_str or str(time_str).strip() == "":
        return ""
    s = str(time_str).strip()
    if not re.match(r"^\d{1,2}:\d{2}$", s):
        print(f"  WARNING: Invalid time '{s}' {context}")
        return s
    return s


def read_sheet(ws, expected_cols, sheet_name):
    """Read a worksheet into a dict keyed by DD/MM/YYYY."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  ERROR: Sheet '{sheet_name}' is empty!")
        return {}

    # Validate headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    for col in expected_cols:
        if col not in headers:
            print(f"  WARNING: Expected column '{col}' not found in '{sheet_name}'. Found: {headers}")

    col_idx = {h: i for i, h in enumerate(headers)}
    data = {}
    warnings = 0

    for row_num, row in enumerate(rows[1:], start=2):
        # Get date
        date_val = row[col_idx.get("date", 0)]
        if date_val is None:
            continue

        # Normalise date to DD/MM/YYYY string
        if isinstance(date_val, datetime):
            date_key = date_val.strftime("%d/%m/%Y")
        else:
            date_key = str(date_val).strip()
            # Handle DD/MM/YYYY format already
            if not re.match(r"\d{2}/\d{2}/\d{4}", date_key):
                print(f"  WARNING: Unexpected date format '{date_key}' at row {row_num} in '{sheet_name}'")
                warnings += 1
                continue

        # Build entry
        entry = {}
        for col_name in expected_cols:
            if col_name == "date":
                continue
            idx = col_idx.get(col_name)
            if idx is not None and idx < len(row):
                val = row[idx]
                entry[col_name] = validate_time(val, f"(row {row_num}, col '{col_name}' in '{sheet_name}')")
            else:
                entry[col_name] = ""

        data[date_key] = entry

    if warnings:
        print(f"  {warnings} warning(s) in sheet '{sheet_name}'")

    return data


def read_hijri(ws):
    """Read Hijri dates from the 'Athan without overwrite' sheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    hijri_data = {}
    hijri_idx = col_idx.get("hijri")
    date_idx  = col_idx.get("date", 0)

    if hijri_idx is None:
        print("  WARNING: No 'hijri' column found in 'Athan without overwrite'")
        return {}

    for row in rows[1:]:
        date_val = row[date_idx]
        if date_val is None:
            continue

        if isinstance(date_val, datetime):
            date_key = date_val.strftime("%d/%m/%Y")
        else:
            date_key = str(date_val).strip()

        hijri_val = row[hijri_idx] if hijri_idx < len(row) else ""
        hijri_data[date_key] = str(hijri_val).strip() if hijri_val else ""

    return hijri_data


def main():
    # Find the xlsx file
    explicit = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = find_xlsx(explicit)

    if not xlsx_path:
        print("ERROR: Could not find masjidbox xlsx file.")
        print("Usage: python xlsx-to-json.py [path/to/masjidbox.xlsx]")
        sys.exit(1)

    print(f"📖 Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Validate sheets exist
    for sheet in [ATHAN_SHEET, IQAMAH_SHEET]:
        if sheet not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet}' not found! Available: {wb.sheetnames}")
            sys.exit(1)

    # Read all data
    print(f"  Reading '{ATHAN_SHEET}'...")
    athan_data = read_sheet(wb[ATHAN_SHEET], ATHAN_COLS, ATHAN_SHEET)
    print(f"    → {len(athan_data)} days loaded")

    print(f"  Reading '{IQAMAH_SHEET}'...")
    iqamah_data = read_sheet(wb[IQAMAH_SHEET], IQAMAH_COLS, IQAMAH_SHEET)
    print(f"    → {len(iqamah_data)} days loaded")

    hijri_data = {}
    if ATHAN_NO_OW in wb.sheetnames:
        print(f"  Reading Hijri dates from '{ATHAN_NO_OW}'...")
        hijri_data = read_hijri(wb[ATHAN_NO_OW])
        print(f"    → {len(hijri_data)} Hijri dates loaded")

    wb.close()

    # ── Merge into final structure ─────────────────────────────────────────
    print("\n🔧 Merging data...")
    days = {}
    all_dates = sorted(set(list(athan_data.keys()) + list(iqamah_data.keys())))

    missing_athan  = 0
    missing_iqamah = 0

    for date_key in all_dates:
        athan  = athan_data.get(date_key, {})
        iqamah = iqamah_data.get(date_key, {})

        if not athan:
            missing_athan += 1
        if not iqamah:
            missing_iqamah += 1

        days[date_key] = {
            "fajr":    athan.get("fajr", ""),
            "sunrise": athan.get("sunrise", ""),
            "dhuhr":   athan.get("dhuhr", ""),
            "asr":     athan.get("asr", ""),
            "maghrib": athan.get("maghrib", ""),
            "isha":    athan.get("isha", ""),
            "jumuah":  athan.get("jumuah", ""),
            "hijri":   hijri_data.get(date_key, ""),
            "iqamah": {
                "fajr":    iqamah.get("fajr", ""),
                "dhuhr":   iqamah.get("dhuhr", ""),
                "asr":     iqamah.get("asr", ""),
                "maghrib": iqamah.get("maghrib", ""),
                "isha":    iqamah.get("isha", ""),
                "jumuah":  iqamah.get("jumuah", ""),
            }
        }

    if missing_athan:
        print(f"  ⚠️  {missing_athan} dates missing athan data")
    if missing_iqamah:
        print(f"  ⚠️  {missing_iqamah} dates missing iqamah data")

    # Detect year from first date
    first_date = all_dates[0] if all_dates else "01/01/2026"
    year = int(first_date.split("/")[2])

    # Build final JSON
    output = {
        "meta": {
            "year": year,
            "generated": datetime.now().isoformat(),
            "source": os.path.basename(xlsx_path),
            "totalDays": len(days),
        },
        "days": days,
    }

    # ── Write output files ─────────────────────────────────────────────────
    script_dir = os.path.dirname(os.path.abspath(__file__))

    for rel_path, label in [
        (THEME_OUTPUT, "WordPress theme"),
        (STATIC_OUTPUT, "Static prototype"),
    ]:
        out_path = os.path.join(script_dir, rel_path)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=None, separators=(",", ":"))

        size_kb = os.path.getsize(out_path) / 1024
        print(f"  ✅ {label}: {out_path} ({size_kb:.1f} KB)")

    # ── Summary ────────────────────────────────────────────────────────────
    print(f"\n✅ Done! {len(days)} days of prayer times generated for {year}.")

    # Quick sanity check: show today's data
    today_key = datetime.now().strftime("%d/%m/%Y")
    if today_key in days:
        d = days[today_key]
        print(f"\n📅 Today ({today_key}):")
        print(f"   Adhan  — Fajr: {d['fajr']}, Dhuhr: {d['dhuhr']}, Asr: {d['asr']}, Maghrib: {d['maghrib']}, Isha: {d['isha']}")
        iq = d['iqamah']
        print(f"   Iqamah — Fajr: {iq['fajr']}, Dhuhr: {iq['dhuhr']}, Asr: {iq['asr']}, Maghrib: {iq['maghrib']}, Isha: {iq['isha']}")
        if d['hijri']:
            print(f"   Hijri  — {d['hijri']}")
    else:
        print(f"\n⚠️  Today's date ({today_key}) not found in data — this is expected if the xlsx is for a different year.")


if __name__ == "__main__":
    main()
